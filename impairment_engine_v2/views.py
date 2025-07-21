import base64
import csv
import io
import logging
import random
from collections import defaultdict
from decimal import Decimal
from io import BytesIO

import numpy as np
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.views.decorators.http import require_http_methods
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from .ecl_computations import ProjectECLProcessor
from .forms import (
    CompanyForm, ProjectForm, BranchMappingForm, BranchMappingBulkForm,
    CBLParametersForm, DataUploadForm, CompanyParametersUpdateForm, LGDRiskFactorForm, LGDRiskFactorValueForm
)
from .matrix_functions import ProjectPDProcessor, IFRS9PDCalculator
from .models import (
    Company, Project, BranchMapping, CBLParameters, LGDRiskFactor, LGDRiskFactorValue, OLSCoefficient
)
from .utils import compute_cumulative_loan_gd, enrich_project_loan_data, compute_final_lgd

logger = logging.getLogger(__name__)


def is_superuser(user):
    return user.is_superuser


@login_required
def home(request):
    """Home page showing companies based on user permissions"""
    if request.user.is_authenticated:
        if request.user.is_superuser:
            company_list = Company.objects.all().order_by('name')
        else:
            company_list = Company.objects.filter(created_by=request.user).order_by('name')

        paginator = Paginator(company_list, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        return render(request, 'impairment_engine/home.html', {'page_obj': page_obj})
    else:
        return redirect('sign_in')


@login_required
def create_company(request):
    """Create a new company with default IFRS9 and CBL parameters"""
    if request.method == 'POST':
        form = CompanyForm(request.POST)
        if form.is_valid():
            company = form.save(commit=False)
            company.created_by = request.user
            company.save()

            # Temporarily create branch mappings
            branches = [
                ('CHIPATA', 'ZW0010001'),
                ('KABWE', 'ZW0010002'),
                ('MANDA HILL', 'ZW0010003'),
                ('CENTRO MALL', 'ZW0010004'),
                ('CAIRO ROAD', 'ZW0010005'),
                ('LONG ACRES', 'ZW0010006'),
                ('KITWE', 'ZW0010007'),
                ('NDOLA', 'ZW0010008'),
                ('LUSAKA', 'ZW0010009'),
            ]
            for name, code in branches:
                BranchMapping.objects.create(
                    company=company,
                    branch_name=name,
                    branch_code=code,
                    is_active=True
                )
            messages.success(request, "Company Added Successfully!")
            return redirect('configure_risk_factors', company_slug=company.slug)
    else:
        form = CompanyForm()
    return render(request, 'impairment_engine/create_company.html', {'form': form})


@login_required
def download_branch_mappings_template(request):
    wb = Workbook()

    # Create the branch mappings Excel sheet
    branches_sheet = wb.active
    branches_sheet.title = "Branch Mappings"
    branches_sheet.append(["Branch Code", "Branch Name"])
    # Add the mappings
    branches_sheet.append(["ZM0010001", "Chipata"])
    branches_sheet.append(["ZM0010002", "Lusaka"])
    branches_sheet.append(["ZM0010003", "Kabwe"])
    branches_sheet.append(["ZM0010004", "Cairo Road"])
    branches_sheet.append(["ZM0010005", "Manda Hill"])
    branches_sheet.append(["ZM0010006", "Long Acres"])
    branches_sheet.append(["ZM0010007", "Ndola"])
    branches_sheet.append(["ZM0010008", "Centro Mall"])
    branches_sheet.append(["ZM0010009", "Kitwe"])

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=branch_mappings_template.xlsx'
    wb.save(response)

    return response


@login_required
def upload_branch_mappings(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug)

    if request.method == 'POST' and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            # Read the Excel file
            wb = load_workbook(excel_file, data_only=True)  # Add data_only to read values not formulas
            if "Branch Mappings" in wb.sheetnames:
                ws = wb["Branch Mappings"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):  # Skip empty rows
                        continue
                    try:
                        branch_code, branch_name = row[:2]
                        BranchMapping.objects.create(
                            branch_code=branch_code,
                            branch_name=branch_name,
                        )
                    except (ValueError, IndexError) as e:
                        messages.warning(request, f"Skipping invalid row in Branch Mappings: {row} - {str(e)}")
                        continue
            messages.success(request, "Branch Mappings Uploaded Successfully!")
            redirect("upload_risk_factors", company_slug=company.slug)
        except Exception as e:
            messages.error(request, f"Error processing excel file: {str(e)}")
            redirect("configure_branch_mappings", company_slug=company.slug)

    return render(request, "impairment_engine/upload_branch_mappings.html", {
        "company_slug": company_slug
    })


@login_required
def upload_risk_factors(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug)

    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            # Read the Excel file
            wb = load_workbook(excel_file, data_only=True)  # Add data_only to read values not formulas

            # Process Risk Factors Sheet
            if "Risk Factors" in wb.sheetnames:
                factors_sheet = wb["Risk Factors"]
                for row in factors_sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):  # Skip empty rows
                        continue
                    try:
                        accessor_key, name, desc = row[:3]  # Get first 3 columns
                        LGDRiskFactor.objects.update_or_create(
                            company=company,
                            accessor_key=accessor_key,
                            defaults={
                                "name": name,
                                "description": desc,
                                "is_active": True,
                            }
                        )
                    except (ValueError, IndexError) as e:
                        messages.warning(request, f"Skipping invalid row in Risk Factors: {row} - {str(e)}")
                        continue

            # Process Risk Factor Values
            if "Risk Factor Values" in wb.sheetnames:
                values_sheet = wb["Risk Factor Values"]
                for row in values_sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):  # Skip empty rows
                        continue
                    try:
                        factor_name, value_name, identifier, lgd_percentage, coefficient = row[:5]
                        factor = LGDRiskFactor.objects.get(company=company, name=factor_name)
                        factor_value, created = LGDRiskFactorValue.objects.update_or_create(
                            factor=factor,
                            name=value_name,
                            defaults={
                                "identifier": identifier,
                                "lgd_percentage": Decimal(str(lgd_percentage)),
                                "is_active": True,
                            }
                        )

                        # Update or create OLSCoefficient
                        OLSCoefficient.objects.update_or_create(
                            company=company,
                            factor_value=factor_value,
                            defaults={"coefficient": Decimal(str(coefficient))}
                        )
                    except LGDRiskFactor.DoesNotExist:
                        messages.warning(request, f"Factor not found: {factor_name} - skipping row")
                        continue
                    except (ValueError, IndexError) as e:
                        messages.warning(request, f"Skipping invalid row in Factor Values: {row} - {str(e)}")
                        continue
                    except Exception as e:
                        messages.warning(request, f"Error processing row: {row} - {str(e)}")
                        continue

            messages.success(request, "Risk Factors successfully imported from excel!")
            return redirect("configure_risk_factors", company_slug=company.slug)

        except Exception as e:
            messages.error(request, f"Error processing excel file: {str(e)}")
            return redirect("configure_risk_factors", company_slug=company.slug)

    return render(request, "impairment_engine/upload_risk_factors.html", {
        "company_slug": company_slug
    })


@login_required
def download_risk_factors_template(request):
    wb = Workbook()

    # Create "Risk Factors" sheet (this becomes the active sheet)
    factors_sheet = wb.active
    factors_sheet.title = "Risk Factors"
    factors_sheet.append(["Accessor Key", "Factor Name", "Description"])

    # Add some example rows
    factors_sheet.append(["client_type", "Client Type", "Type of client"])
    factors_sheet.append(["collateral_type", "Collateral Type", "Type of collateral"])

    # Create "Factor Values" sheet
    values_sheet = wb.create_sheet("Risk Factor Values")
    values_sheet.append(["Factor Name", "Value Name", "Unique Identifier", "LGD Percentage", "Coefficient"])

    # Add some example rows
    values_sheet.append(["Client Type", "Individual", "1", "45.00", "0.123456"])
    values_sheet.append(["Client Type", "Corporate", "2", "35.00", "0.098765"])
    values_sheet.append(["Collateral Type", "Real Estate", "1", "30.00", "0.080000"])

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=risk_factor_template.xlsx'
    wb.save(response)

    return response


@login_required
def configure_risk_factors(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug)

    if request.method == 'POST':
        factor_form = LGDRiskFactorForm(request.POST)
        if factor_form.is_valid():
            factor = factor_form.save(commit=False)
            factor.company = company
            factor.save()
            messages.success(request, f"Factor '{factor.name}' added.")
            return redirect('configure_risk_factors', company_slug=company_slug)
    else:
        factor_form = LGDRiskFactorForm()

    risk_factors = company.risk_factors.prefetch_related('values')

    return render(request, 'impairment_engine/configure_risk_factors.html', {
        'company': company,
        'factor_form': factor_form,
        'risk_factors': risk_factors
    })


@login_required
def add_risk_factor_value(request, company_slug, factor_id):
    factor = get_object_or_404(LGDRiskFactor, id=factor_id)
    company = get_object_or_404(Company, slug=company_slug)

    if request.method == 'POST':
        form = LGDRiskFactorValueForm(request.POST)
        if form.is_valid():
            value = form.save(commit=False)
            value.factor = factor
            value.save()

            # Check if user submitted OLS coefficient
            coefficient = form.cleaned_coefficient
            if coefficient is not None:
                OLSCoefficient.objects.create(
                    company=company,
                    factor_value=value,
                    coefficient=coefficient,
                )
            messages.success(request, f"Value '{value.name}' added to {factor.name}.")
        else:
            messages.warning(request, "Invalid form provided. Please try again.")
    return redirect('configure_risk_factors', company_slug=factor.company.slug)


@login_required
def company_detail(request, company_slug):
    """Company detail view showing overview and quick stats"""
    company = get_object_or_404(Company, slug=company_slug)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to access this company.")
        return redirect('home')

    # Get company statistics
    projects = company.projects.all()
    active_projects = projects.filter(status__in=['setup', 'data_upload', 'processing', 'validation'])
    completed_projects = projects.filter(status='completed')
    branch_mappings = company.branch_mappings.filter(is_active=True)

    context = {
        'company': company,
        'projects': projects[:5],  # Show latest 5 projects
        'active_projects_count': active_projects.count(),
        'completed_projects_count': completed_projects.count(),
        'branch_mappings_count': branch_mappings.count(),
        'total_projects': projects.count(),
    }

    return render(request, 'impairment/company_detail.html', context)


@login_required
def company_projects(request, company_slug):
    """List all projects for a company"""
    company = get_object_or_404(Company, slug=company_slug)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to access this company.")
        return redirect('home')

    # Check if L.G.D Factors have been set
    risk_factors = company.risk_factors.prefetch_related('values')

    if risk_factors.count() == 0:
        return redirect('configure_risk_factors', company_slug=company.slug)

    projects_list = company.projects.all().order_by('-created_at')
    paginator = Paginator(projects_list, 15)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'company': company,
        'page_obj': page_obj,
    }

    return render(request, 'impairment_engine/company_projects.html', context)


@login_required
def create_project(request, company_slug):
    """Create a new project within a company"""
    company = get_object_or_404(Company, slug=company_slug)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to create projects for this company.")
        return redirect('index')

    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.company = company
            project.created_by = request.user
            project.save()
            messages.success(request, "Project Created Successfully!")
            return HttpResponseRedirect(reverse('upload_wizard', args=[company_slug, project.slug]))
    else:
        form = ProjectForm()

    return render(request, 'impairment_engine/create_project.html', {'company': company, 'form': form})


@login_required
def dashboard(request, company_slug, project_slug):
    """Project detail view showing status and progress"""
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    data = pd.DataFrame(project.loan_data)

    # Data Transformation
    usd_rate = 23.02
    # data["exposure"] = np.where(data["currency"] == "USD", data["exposure"] * usd_rate, data["exposure"])
    # data["total_ecl"] = np.where(data["currency"] == "USD", data["total_ecl"] * usd_rate, data["total_ecl"])

    # Calculate metrics by loan stage
    stage_metrics = data.groupby("loan_stage").agg(
        total_ecl=('total_ecl', 'sum'),
        total_exposure=('exposure', 'sum'),
        loan_count=('account_number', 'count')
    ).reset_index()

    # Add percentage of total exposure for each stage
    total_exposure = stage_metrics['total_exposure'].sum()
    stage_metrics['exposure_pct'] = (stage_metrics['total_exposure'] / total_exposure * 100).round(2)



    # Calculate metrics by loan type
    loan_type_metrics = data.groupby('loan_type').agg(
        total_ecl=('total_ecl', 'sum'),
        total_exposure=('exposure', 'sum'),
        loan_count=('account_number', 'count')
    ).reset_index()

    # Add percentage of total exposure for each loan type
    loan_type_metrics['exposure_pct'] = (loan_type_metrics['total_exposure'] / total_exposure * 100).round(2)
    loan_type_metrics['type_pct'] = (loan_type_metrics['loan_count'] / len(data) * 100).round(2)

    # Calculate overall metrics
    overall_metrics = {
        'loan_types': len(loan_type_metrics),
        'total_loans': len(data),
        'total_exposure': data['exposure'].sum(),
        'total_ecl': data['total_ecl'].sum(),
        'ecl_ratio': (data['total_ecl'].sum() / data['exposure'].sum() * 100).round(2) if data[
                                                                                              'exposure'].sum() > 0 else 0,
    }

    # Convert DataFrames to dict for template
    # stage_metrics_dict = stage_metrics.to_dict('records')
    loan_type_metrics_dict = loan_type_metrics.to_dict('records')

    stage_metrics_dict = {
        metric['loan_stage']: metric
        for metric in stage_metrics.to_dict('records')
    }

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to access this project.")
        return redirect('home')

    print(stage_metrics_dict)

    context = {
        'company': company,
        'project': project,
        'stage_metrics': stage_metrics_dict,
        'loan_type_metrics': loan_type_metrics_dict,
        'overall_metrics': overall_metrics,
    }

    return render(request, 'impairment_engine/project_dashboard.html', context)


@login_required
def data_upload_wizard(request, company_slug, project_slug):
    # Fetch company and project details
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to view this project dashboard.")
        return redirect('index')

    # Check if data upload not already processed
    if project.loan_report_uploaded and project.arrears_report_uploaded and project.status != "":
        messages.error(request, "Project data has already been uploaded. Navigating to project dashboard.")
        return redirect('current_loan_book', company_slug=company_slug, project_slug=project_slug, stage='stage_1')

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        print(f"INFO: Excel file received: {excel_file}")
        print(f"INFO: Excel file name: {excel_file.name if excel_file else 'None'}")

        if not excel_file:
            print(f"ERROR: No excel file provided")
            messages.error(request, "Please select an excel file for upload.")
            return render(request, 'impairment_engine/data_upload_wizard.html', {
                'company_slug': company_slug,
                'project_slug': project_slug,
                'company': company,  # Added this
                'project': project,  # Added this
                'step': 1
            })

        try:
            print(f"INFO: Attempting to read Excel file")
            # Reset file pointer to beginning
            excel_file.seek(0)
            xlsx = pd.ExcelFile(excel_file)
            print(f"INFO: Excel file loaded successfully")
            print(f"INFO: Sheet names: {xlsx.sheet_names}")

            # Read file content and encode as base64 for session storage
            excel_file.seek(0)  # Reset file pointer again
            file_content = excel_file.read()
            file_content_b64 = base64.b64encode(file_content).decode('utf-8')

            # Store data in session
            request.session['upload_data'] = {
                'file_name': excel_file.name,
                'sheet_names': xlsx.sheet_names,
                'file_content_b64': file_content_b64  # Store as base64 string
            }
            request.session.modified = True
            return redirect('process_sheets', company_slug=company_slug, project_slug=project_slug)

        except Exception as e:
            import traceback
            messages.error(request, f"Error reading excel file: {str(e)}")
            return render(request, 'impairment_engine/data_upload_wizard.html', {
                'company_slug': company_slug,
                'project_slug': project_slug,
                'company': company,
                'project': project,
                'step': 1
            })

    return render(request, 'impairment_engine/data_upload_wizard.html', {
        'company_slug': company_slug,
        'project_slug': project_slug,
        'company': company,  # Added this
        'project': project,  # Added this
        'step': 1
    })


@login_required
def process_sheet_selection(request, company_slug, project_slug):
    if 'upload_data' in request.session:
        print(f"DEBUG: upload_data contents: {list(request.session['upload_data'].keys())}")

    if 'upload_data' not in request.session:
        print(f"DEBUG: No upload_data in session, redirecting to step 1 of the data upload component.")
        messages.error(request, "Please select an excel file for upload.")
        return redirect('upload_wizard', company_slug=company_slug, project_slug=project_slug)

    if request.method == 'POST':
        loan_sheet = request.POST.get('loan_sheet')
        arrears_sheet = request.POST.get('arrears_sheet')
        deposit_listing_sheet = request.POST.get('deposit_listing_sheet')
        print(f"DEBUG: POST data - loan_sheet: {loan_sheet}, arrears_sheet: {arrears_sheet}, deposit_listing_sheet: {deposit_listing_sheet}")

        if not loan_sheet or not arrears_sheet or not deposit_listing_sheet:
            messages.error(request, "Please select loans, arrears and deposit listing sheets.")
            return render(request, 'impairment_engine/data_upload_wizard.html', {
                'company_slug': company_slug,
                'project_slug': project_slug,
                'sheet_names': request.session['upload_data']['sheet_names'],
                'step': 2
            })

        # Store sheet selections in session
        request.session['upload_data'].update({
            'loan_sheet': loan_sheet,
            'arrears_sheet': arrears_sheet,
            'deposit_listing_sheet': deposit_listing_sheet
        })
        request.session.modified = True
        print(f"DEBUG: Updated session with sheet selections")

        return redirect('process_mapping', company_slug=company_slug, project_slug=project_slug)

    print(f"DEBUG: Rendering step 2 template")
    print(f"DEBUG: Available sheet names: {request.session['upload_data']['sheet_names']}")
    return render(request, 'impairment_engine/data_upload_wizard.html', {
        'company_slug': company_slug,
        'project_slug': project_slug,
        'sheet_names': request.session['upload_data']['sheet_names'],
        'step': 2
    })


@login_required
def process_column_mapping(request, company_slug, project_slug):
    print(f"DEBUG: process_column_mapping called")
    ARREARS_BUCKETS = [
        ('00-07 DAYS', '0_7_days', 0, 7),
        ('08-14 DAYS', '8_14_days', 8, 14),
        ('15-30 DAYS', '15_30_days', 15, 30),
        ('31-60 DAYS', '31_60_days', 31, 60),
        ('61-90 DAYS', '61_90_days', 61, 90),
        ('91-120 DAYS', '91_120_days', 91, 120),
        ('121-150 DAYS', '121_150_days', 121, 150),
        ('151-180 DAYS', '151_180_days', 151, 180),
        ('181-360 DAYS', '181_360_days', 181, 360),
        ('OVER 360 DAYS', 'over_360_days', 361, 365)
    ]

    if 'upload_data' not in request.session or 'loan_sheet' not in request.session['upload_data']:
        print(f"DEBUG: Missing session data, redirecting to step 1 of the data upload component.")
        messages.error(request, "Please complete previous steps first")
        return redirect('upload_wizard', company_slug=company_slug, project_slug=project_slug)

    try:
        # Decode the base64 file content
        file_content_b64 = request.session['upload_data']['file_content_b64']
        file_content = base64.b64decode(file_content_b64.encode('utf-8'))
        xls = pd.ExcelFile(BytesIO(file_content))

        loan_df = pd.read_excel(xls, sheet_name=request.session['upload_data']['loan_sheet'], nrows=1)
        arrears_df = pd.read_excel(xls, sheet_name=request.session['upload_data']['arrears_sheet'], nrows=1)
        deposit_listing_df = pd.read_excel(xls, sheet_name=request.session['upload_data']['deposit_listing_sheet'], nrows=1)

        loan_columns = loan_df.columns.tolist()
        arrears_columns = arrears_df.columns.tolist()
        deposit_listing_columns = deposit_listing_df.columns.tolist()
        print(f"DEBUG: Loan columns: {loan_columns}")
        print(f"DEBUG: Arrears columns: {arrears_columns}")
        print(f"DEBUG: Deposit listing columns: {deposit_listing_columns}")

        # Check if bucket columns are present in arrears data
        bucket_columns_found = []
        for bucket_name, _, _, _ in ARREARS_BUCKETS:
            if bucket_name in arrears_columns:
                bucket_columns_found.append(bucket_name)

        bucket_columns_present = len(bucket_columns_found) > 0
        request.session['upload_data']['has_bucket_columns'] = bucket_columns_present
        request.session['upload_data']['bucket_columns_found'] = bucket_columns_found

        print(f"DEBUG: Bucket columns found: {bucket_columns_found}")
        print(f"DEBUG: Has bucket columns: {bucket_columns_present}")

        request.session.modified = True

    except Exception as e:
        print(f"DEBUG: Error reading Excel sheets: {str(e)}")
        import traceback
        print(f"DEBUG: Full traceback: {traceback.format_exc()}")
        messages.error(request, f"Error reading Excel sheets: {str(e)}")
        return redirect('upload_wizard', company_slug=company_slug, project_slug=project_slug)

    # Define field mappings with descriptions
    loan_field_config = [
        ('account_number', 'Account Number', True),
        # ('branch', 'Branch', True), # Prioritizing Deposit Listing
        # ('client_name', 'Client Name', True), # Prioritizing Deposit Listing
        ('loan_type', 'Loan Type', True),
        ('loan_amount', 'Loan Amount', True),
        ('currency', 'Currency', True),
        ('opening_date', 'Opening Date', True),
        ('maturity_date', 'Maturity Date', True),
        ('installment_amount', 'Installment Amount', True),
        ('capital_balance', 'Capital Balance', True),
        ('interest_rate', 'Interest Rate', True),
        ('loan_tenor', 'Loan Tenor (Days)', False),  # Can be computed
        ('days_to_maturity', 'Days to Maturity', False),  # Can be computed
    ]

    deposit_listing_field_config = [
        ('account_number', 'Account Number', True),
        ('branch', 'Branch', True),
        ('client_name', 'Client Name', True),
        ('sector', 'Sector', True),
    ]
    # Conditional arrears field config based on bucket presence
    if bucket_columns_present:
        arrears_field_config = [
            ('account_number', 'Account Number', True),
            ('currency', 'Currency', False),  # Optional since might come from loan sheet
            ('capital_balance', 'Capital Balance', False),  # Optional since might come from loan sheet
        ]
    else:
        arrears_field_config = [
            ('account_number', 'Account Number', True),
            ('currency', 'Currency', True),
            ('capital_balance', 'Capital Balance', True),
            ('arrears_amount', 'Arrears Amount', True),
            ('exposure', 'Exposure Amount', False),  # Can be computed
            ('days_past_due', 'Days Past Due', True),
        ]

    if request.method == 'POST':
        print(f"DEBUG: Processing column mapping POST request")
        # Process the mappings
        mappings = {
            'loan_mappings': {},
            'arrears_mappings': {},
            'deposit_listing_mappings': {},
        }

        # Process loan mappings
        for field_name, description, is_required in loan_field_config:
            selected_column = request.POST.get(f'loan_{field_name}')
            if selected_column:
                mappings['loan_mappings'][field_name] = selected_column

        # Process arrears mappings
        for field_name, description, is_required in arrears_field_config:
            selected_column = request.POST.get(f'arrears_{field_name}')
            if selected_column:
                mappings['arrears_mappings'][field_name] = selected_column

        # Process arrears mappings
        for field_name, description, is_required in deposit_listing_field_config:
            selected_column = request.POST.get(f'deposit_listing_{field_name}')
            if selected_column:
                mappings['deposit_listing_mappings'][field_name] = selected_column

        print(f"DEBUG: Mappings created: {mappings}")

        # Store mappings in session
        request.session['upload_data']['mappings'] = mappings
        request.session.modified = True
        return redirect('finalize_upload', company_slug=company_slug, project_slug=project_slug)

    return render(request, 'impairment_engine/data_upload_wizard.html', {
        'company_slug': company_slug,
        'project_slug': project_slug,
        'loan_columns': loan_columns,
        'arrears_columns': arrears_columns,
        'deposit_listing_columns': deposit_listing_columns,
        'loan_field_config': loan_field_config,
        'arrears_field_config': arrears_field_config,
        'deposit_listing_field_config': deposit_listing_field_config,
        'step': 3,
        'ARREARS_BUCKETS': ARREARS_BUCKETS,
        'has_bucket_columns': bucket_columns_present,
        'bucket_columns_found': bucket_columns_found,
    })


@login_required
def finalize_data_upload_v2(request, company_slug, project_slug):
    print(f"DEBUG: Finalize_data_upload called")
    if 'upload_data' not in request.session or 'mappings' not in request.session['upload_data']:
        print(f"DEBUG: Missing session data for finalize")
        messages.error(request, "Please complete all steps first")
        return redirect('current_loan_book', company_slug=company_slug, project_slug=project_slug, stage='stage_1')

    try:
        project = Project.objects.get(slug=project_slug, company__slug=company_slug)
        company = Company.objects.get(slug=company_slug)
        upload_data = request.session['upload_data']

        # Temporarily define rate for USD
        rate = 13.7031

        # Add this check at the start of finalize_data_upload_v2
        print(
            f"DEBUG: Company Stage Thresholds - Stage 1: {company.stage_1_threshold_days}, Stage 2: {company.stage_2_threshold_days}")

        # Decode the base64 file content
        file_content_b64 = upload_data['file_content_b64']
        file_content = base64.b64decode(file_content_b64.encode('utf-8'))
        xls = pd.ExcelFile(BytesIO(file_content))

        # Update status to data upload processing
        project.status = 'processing'

        # Process loan data
        print(f"DEBUG: Processing loan data")
        loan_df = pd.read_excel(xls, sheet_name=upload_data['loan_sheet'])
        print(f"Loans columns currently {loan_df.columns}")

        # Invert the mapping to go from source_column -> target_column
        loan_mapping_inverted = {v: k for k, v in upload_data['mappings']['loan_mappings'].items()}
        print(f"DEBUG: Loan mapping inverted: {loan_mapping_inverted}")
        loan_df = loan_df.rename(columns=loan_mapping_inverted)

        # Keep only mapped columns for loan data
        loan_mapped_columns = list(loan_mapping_inverted.values())
        loan_df = loan_df[loan_mapped_columns]
        print(f"Loans columns after filtering: {loan_df.columns}")

        # Process arrears data
        print(f"DEBUG: Processing arrears data")
        arrears_df = pd.read_excel(xls, sheet_name=upload_data['arrears_sheet'])
        print(f"Arrears columns currently {arrears_df.columns}")

        # Define arrears buckets at module level for consistency
        ARREARS_BUCKETS = [
            ('00-07 DAYS', '0_7_days', 0, 7),
            ('08-14 DAYS', '8_14_days', 8, 14),
            ('15-30 DAYS', '15_30_days', 15, 30),
            ('31-60 DAYS', '31_60_days', 31, 60),
            ('61-90 DAYS', '61_90_days', 61, 90),
            ('91-120 DAYS', '91_120_days', 91, 120),
            ('121-150 DAYS', '121_150_days', 121, 150),
            ('151-180 DAYS', '151_180_days', 151, 180),
            ('181-360 DAYS', '181_360_days', 181, 360),
            ('OVER 360 DAYS', 'over_360_days', 361, 365)
        ]

        # Process bucket-based arrears if bucket columns are present
        if upload_data.get('has_bucket_columns', False):
            print(f"DEBUG: Processing bucket-based arrears data")

            # Initialize arrears columns
            arrears_df['arrears_amount'] = 0.0
            arrears_df['days_past_due'] = 0

            # Process each row to find which bucket has the arrears amount
            for idx, row in arrears_df.iterrows():
                total_arrears = 0.0
                max_dpd = 0  # Track the highest DPD bucket with arrears

                for bucket_name, field_name, min_days, max_days in ARREARS_BUCKETS:
                    bucket_value = 0.0

                    # Check if this exact bucket column exists and has a value
                    if bucket_name in arrears_df.columns:
                        cell_value = row[bucket_name]

                        # Handle different representations of empty/zero values
                        if pd.isna(cell_value) or cell_value == '-' or cell_value == '' or cell_value == 0:
                            bucket_value = 0.0
                        else:
                            try:
                                # Handle string numbers with commas
                                if isinstance(cell_value, str):
                                    cell_value = cell_value.replace(',', '').replace(' ', '')
                                bucket_value = float(cell_value)
                            except (ValueError, TypeError):
                                bucket_value = 0.0

                        if bucket_value > 0:
                            total_arrears += (bucket_value / rate) # Divide by the rate for accurate USD Reporting
                            # Temporarily set max_dpd to a random number
                            # max_dpd = max(max_dpd, max_days)
                            max_dpd = random.randint(max_days, max_days + 1)

                arrears_df.at[idx, 'arrears_amount'] = round(total_arrears, 2)
                arrears_df.at[idx, 'days_past_due'] = max_dpd

            print(
                f"DEBUG: Processed {len(arrears_df[arrears_df['arrears_amount'] > 0])} accounts with arrears from bucket format")

            # Invert the mapping for arrears (only for account_number and other mapped fields)
            arrears_mapping_inverted = {v: k for k, v in upload_data['mappings']['arrears_mappings'].items()}
            print(f"DEBUG: Arrears mapping inverted: {arrears_mapping_inverted}")

            # Rename only the mapped columns
            arrears_df = arrears_df.rename(columns=arrears_mapping_inverted)

            # Keep mapped columns plus our computed ones
            arrears_mapped_columns = list(arrears_mapping_inverted.values()) + ['arrears_amount', 'days_past_due']
            # Filter to only columns that exist
            arrears_mapped_columns = [col for col in arrears_mapped_columns if col in arrears_df.columns]
            arrears_df = arrears_df[arrears_mapped_columns]

        else:
            # Traditional processing for non-bucket data
            print(f"DEBUG: Processing traditional arrears data")
            arrears_mapping_inverted = {v: k for k, v in upload_data['mappings']['arrears_mappings'].items()}
            print(f"DEBUG: Arrears mapping inverted: {arrears_mapping_inverted}")
            arrears_df = arrears_df.rename(columns=arrears_mapping_inverted)

            # Keep only mapped columns for arrears data
            arrears_mapped_columns = list(arrears_mapping_inverted.values())
            arrears_df = arrears_df[arrears_mapped_columns]

        print(f"Arrears columns after filtering: {arrears_df.columns}")

        # Process deposit listing data
        print(f"DEBUG: Processing deposit listing data")
        deposit_listing_df = pd.read_excel(xls, sheet_name=upload_data['deposit_listing_sheet'])
        print(f"Deposit Listing columns currently {deposit_listing_df.columns}")

        # Invert the mapping to go from source_column -> target_column
        deposit_listing_mapping_inverted = {v: k for k, v in upload_data['mappings']['deposit_listing_mappings'].items()}
        print(f"DEBUG: Deposit Listing inverted: {deposit_listing_mapping_inverted}")
        deposit_listing_df = deposit_listing_df.rename(columns=deposit_listing_mapping_inverted)

        # Keep only mapped columns for loan data
        deposit_listing_mapped_columns = list(deposit_listing_mapping_inverted.values())
        deposit_listing_df = deposit_listing_df[deposit_listing_mapped_columns]
        print(f"Loans columns after filtering: {deposit_listing_df.columns}")

        # Verify the account_number column exists before merging
        print(f"DEBUG: 'account_number' in loan_df: {'account_number' in loan_df.columns}")
        print(f"DEBUG: 'account_number' in arrears_df: {'account_number' in arrears_df.columns}")

        # Merge loan and arrears data on account_number
        print(f"DEBUG: Merging loan and arrears data")
        merged_df = loan_df.merge(
            arrears_df,
            on='account_number',
            how='left',  # Keep all loans, even those without arrears
            suffixes=('', '_arrears')
        )

        # Merge deposit listing as well
        merged_df = merged_df.merge(
            deposit_listing_df,
            on='account_number',
            how='left',
            suffixes=('', '_listing')
        )

        # Handle missing arrears data, set defaults for accounts not in arrears
        merged_df['days_past_due'] = merged_df['days_past_due'].fillna(0)
        merged_df['arrears_amount'] = merged_df['arrears_amount'].fillna(0)

        # Handle capital balance conflicts (prioritize loan data)
        if 'capital_balance_arrears' in merged_df.columns:
            merged_df['capital_balance'] = merged_df['capital_balance'].fillna(merged_df['capital_balance_arrears'])
            merged_df = merged_df.drop('capital_balance_arrears', axis=1)

        # Handle currency conflicts (prioritize loan data)
        if 'currency_arrears' in merged_df.columns:
            merged_df['currency'] = merged_df['currency'].fillna(merged_df['currency_arrears'])
            merged_df = merged_df.drop('currency_arrears', axis=1)

        """ Calculate computed fields """
        print(f"DEBUG: Calculating computed fields")

        # Calculate exposure
        merged_df["exposure"] = round((merged_df['capital_balance'] + merged_df['arrears_amount']), 2)

        # Calculate loan_tenor in months if dates are available
        if 'opening_date' in merged_df.columns and 'maturity_date' in merged_df.columns:
            # Add temporary date adjust of 1 year
            merged_df['opening_date'] = pd.to_datetime(merged_df['opening_date'], errors='coerce') + pd.DateOffset(years=1)
            merged_df['maturity_date'] = pd.to_datetime(merged_df['maturity_date'], errors='coerce') + pd.DateOffset(years=1)

            # Calculate loan tenor in months (approximate using 30.44 days per month)
            loan_tenor_days = (merged_df['maturity_date'] - merged_df['opening_date']).dt.days
            merged_df['loan_tenor'] = (loan_tenor_days / 30.44).round().astype('Int64')  # Round to nearest month

        # Calculate days_to_maturity (set to 0 if past maturity date)
        if 'maturity_date' in merged_df.columns:
            today = pd.Timestamp.now().normalize()
            days_to_maturity = (merged_df['maturity_date'] - today).dt.days
            merged_df['days_to_maturity'] = days_to_maturity.where(days_to_maturity >= 0, 0)

        # Add loan stage based on days past due
        def get_loan_stage(dpd):
            print(f"DEBUG: Calculating stage for DPD: {dpd} (type: {type(dpd)})")
            if pd.isna(dpd) or dpd == 0:
                return 'stage_1' # Current Performing Loans fall under stage 1
            elif dpd <= company.stage_1_threshold_days:
                return 'stage_1'
            elif dpd <= company.stage_2_threshold_days:
                return 'stage_2'
            else:
                return 'stage_3'

        merged_df['loan_stage'] = merged_df['days_past_due'].apply(get_loan_stage)

        if 'model_pd' not in merged_df.columns:
            merged_df['model_pd'] = None

        # Randomly assign PDs for loan stage
        def generate_random_pd(loan_stage):
            if loan_stage == 'stage_1':
                # Stage 1 Loans low PD between 1 and 6%
                return round(random.uniform(0.01, 0.06), 8)
            elif loan_stage == 'stage_2':
                return round(random.uniform(0.075, 0.15), 8)
            else:
                return round(random.uniform(0.16, 0.35), 8)

        # only fill in the random PD if not provided for the loan
        if merged_df['model_pd'].isna().any():
            mask = merged_df['model_pd'].isna()
            merged_df.loc[mask, 'model_pd'] = merged_df.loc[mask, 'model_pd'].apply(generate_random_pd)
        elif (merged_df['model_pd'] == 0).any():
            # Alternatively, if existing PDs are 0 (rather than NA), treat them as missing
            mask = merged_df['model_pd'] == 0
            merged_df.loc[mask, 'model_pd'] = merged_df.loc[mask, 'loan_stage'].apply(generate_random_pd)

        # Define the expected final columns
        expected_columns = [
            'client_name', 'branch', 'sector', 'account_number', 'loan_type', 'opening_date',
            'maturity_date', 'currency', 'loan_amount', 'capital_balance', 'interest_rate',
            'installment_amount', 'arrears_amount', 'days_past_due', 'exposure',
            'loan_tenor', 'days_to_maturity', 'loan_stage', 'model_pd'
        ]

        # Filter to only keep expected columns that exist in the dataframe
        final_columns = [col for col in expected_columns if col in merged_df.columns]
        merged_df = merged_df[final_columns]

        # Handle branch mapping on the branch column
        def get_branch_name(branch_code):
            branch = BranchMapping.objects.get(branch_code=branch_code, company=company)
            if branch is not None:
                return branch.branch_name
            return branch_code

        merged_df['branch'] = merged_df['branch'].apply(get_branch_name)

        # Ensure all columns are JSON serializable
        for col in merged_df.columns:
            if pd.api.types.is_numeric_dtype(merged_df[col]):
                merged_df[col] = merged_df[col].apply(lambda x: float(x) if pd.notnull(x) else 0)
            elif pd.api.types.is_datetime64_any_dtype(merged_df[col]):
                merged_df[col] = merged_df[col].dt.strftime('%Y-%m-%d')
            elif merged_df[col].dtype == 'object':
                merged_df[col] = merged_df[col].astype(str)

        # Convert to dictionary format and store
        loan_data = merged_df.to_dict(orient='records')

        # Store the merged data
        project.status = 'completed'
        project.loan_data = loan_data
        project.loan_report_uploaded = True
        project.arrears_report_uploaded = True

        # Add metadata about the upload
        project.upload_metadata = {
            'total_accounts': len(loan_data),
            'accounts_with_arrears': len(merged_df[merged_df['days_past_due'] > 0]),
            'accounts_current': len(merged_df[merged_df['days_past_due'] == 0]),
            'upload_date': pd.Timestamp.now().isoformat(),
            'total_exposure': float(merged_df['exposure'].sum()),
            'total_arrears': float(merged_df['arrears_amount'].sum()),
            'status_breakdown': merged_df['loan_stage'].value_counts().to_dict(),
            'processing_type': 'bucket_based' if upload_data.get('has_bucket_columns', False) else 'traditional'
        }

        project.save()

        # Clear session data
        if 'upload_data' in request.session:
            del request.session['upload_data']

        messages.success(request, f"Data uploaded successfully! Processed {len(loan_data)} loan accounts.")
        return redirect('project_dashboard', company_slug=company_slug, project_slug=project_slug)

    except Exception as e:
        print(f"DEBUG: Error in finalize_data_upload: {str(e)}")
        import traceback
        print(f"DEBUG: Full traceback: {traceback.format_exc()}")
        messages.error(request, f"Error processing data: {str(e)}")
        return redirect('upload_wizard', company_slug=company_slug, project_slug=project_slug)


@login_required
def current_loanbook(request, company_slug, project_slug, stage):
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    data = pd.DataFrame(project.loan_data)
    filtered_loans  = data[data['loan_stage'] == stage]
    loans_list = filtered_loans.to_dict(orient='records')

    paginator = Paginator(loans_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Define columns to render
    columns = [
        'account_number',
        'client_name',
        'loan_type',
        'currency',
        'loan_amount',
        'opening_date',
        'maturity_date',
        'loan_tenor'
    ]

    staging_title = ''
    if stage == 'stage_1':
        staging_title = 'Stage 1 Loans - Performing'
    elif stage == 'stage_2':
        staging_title = 'Stage 2 Loans - Significant Increase in Credit Risk'
    elif stage == 'stage_3':
        staging_title = 'Stage 3 Loans - Non-Performing (Defaulted)'

    context = {
        'company': company,
        'project': project,
        'columns': columns,
        'page_obj': page_obj,
        'staging': staging_title
    }

    return render(request, 'impairment_engine/current_loanbook.html', context)


@login_required
def current_cbl(request, company_slug, project_slug):
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    data = pd.DataFrame(project.loan_data)
    # Data transformations to match rates
    usd_rate = 13.7031
    # Loan Amount
    data["loan_amount"] = np.where(data["loan_amount"] == "USD", data["loan_amount"] * usd_rate, data["loan_amount"])
    data["arrears_amount"] = np.where(data["arrears_amount"] == "USD", data["arrears_amount"] * usd_rate, data["arrears_amount"])
    data["exposure"] = np.where(data["exposure"] == "USD", data["exposure"] * usd_rate, data["exposure"])
    data["interest_rate"] = data["interest_rate"].astype(float).round(0).astype(int).astype(str) + "%"

    # Ensure each row is a dict (not Series or string)
    data_dicts = data.to_dict(orient="records")
    paginator = Paginator(data_dicts, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Define columns to render
    columns = [
        'account_number',
        'client_name',
        'loan_type',
        'sector',
        'branch',
        'interest_rate',
        'currency',
        'capital_balance',
        'arrears_amount',
        'exposure'
    ]

    context = {
        'company': company,
        'project': project,
        'columns': columns,
        'page_obj': page_obj
    }

    return render(request, 'impairment_engine/cbl.html', context)


@login_required
def current_exposure(request, company_slug, project_slug):
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    data = pd.DataFrame(project.loan_data)
    # Only take the loans with exposure
    data = data[data["exposure"] > 0]

    # Data transformations to match rates
    usd_rate = 13.7031
    data["loan_amount"] = np.where(data["loan_amount"] == "USD", round(data["loan_amount"] * usd_rate, 2), data["loan_amount"])
    data["arrears_amount"] = np.where(data["arrears_amount"] == "USD", round(data["arrears_amount"] * usd_rate, 2), data["arrears_amount"])
    data["exposure"] = np.where(data["exposure"] == "USD", round(data["exposure"] * usd_rate, 2), data["exposure"])
    data["interest_rate"] = data["interest_rate"].astype(float).round(0).astype(int).astype(str) + "%"

    data["net_disbursement"] = np.where(data["loan_type"] == "Micro Lease Loan", round(data["loan_amount"] * 0.7, 2), data["loan_amount"])
    data["gross_disbursement"] = data["loan_amount"]

    # Ensure each row is a dict (not Series or string)
    data_dicts = data.to_dict(orient="records")
    paginator = Paginator(data_dicts, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Define columns to render
    columns = [
        'account_number',
        'client_name',
        'loan_type',
        'interest_rate',
        'currency',
        'capital_balance',
        'arrears_amount',
        'net_disbursement',
        'gross_disbursement',
        'installment_amount',
        'opening_date',
        'maturity_date',
        'exposure'
    ]

    context = {
        'company': company,
        'project': project,
        'columns': columns,
        'page_obj': page_obj
    }

    return render(request, 'impairment_engine/current_ead.html', context)


@login_required
def compute_project_lgd(request, company_slug, project_slug):
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)
    updated_loan_data = []

    # transform the loan data
    loan_data_df = enrich_project_loan_data(project)

    # First pass: compute all cumulative GDs and store them
    cumulative_gds = []
    loan_dicts = []

    for index, loan in loan_data_df.iterrows():
        loan_dict = loan.to_dict()
        try:
            cumulative_gd = compute_cumulative_loan_gd(company, loan_dict)
            loan_dict["cumulative_gd"] = float(cumulative_gd)
            cumulative_gds.append(float(cumulative_gd))
        except Exception as e:
            print(e)
            loan_dict["cumulative_gd"] = None
            loan_dict["lgd_error"] = str(e)

        loan_dicts.append(loan_dict)

    # Calculate count of records per cumulative GD (rounded to 6 decimals)
    gd_counts = defaultdict(int)
    for gd in cumulative_gds:
        rounded_gd = round(gd, 6)
        gd_counts[rounded_gd] += 1

        # Second pass: compute final LGD for each loan using the new method
        for loan_dict in loan_dicts:
            if loan_dict.get("cumulative_gd") is not None:
                rounded_gd = round(loan_dict["cumulative_gd"], 6)
                count = gd_counts.get(rounded_gd, 1)

                # Use the dedicated method for final LGD computation
                loan_dict["computed_lgd"] = compute_final_lgd(
                    cumulative_gd=loan_dict["cumulative_gd"],
                    count=count
                )

    project.loan_data = loan_dicts
    project.save()

    return redirect("current_loss_given_default", company_slug=company_slug, project_slug=project_slug)

@login_required
def current_loss_given_default(request, company_slug, project_slug):
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    data = pd.DataFrame(project.loan_data)
    # Only take the loans with exposure
    data = data[data["exposure"] > 0]

    # Data transformations to match rates
    usd_rate = 13.7031
    data["loan_amount"] = np.where(data["loan_amount"] == "USD", round(data["loan_amount"] * usd_rate, 2), data["loan_amount"])
    data["arrears_amount"] = np.where(data["arrears_amount"] == "USD", round(data["arrears_amount"] * usd_rate, 2), data["arrears_amount"])
    data["exposure"] = np.where(data["exposure"] == "USD", round(data["exposure"] * usd_rate, 2), data["exposure"])
    data["interest_rate"] = data["interest_rate"].astype(float).round(0).astype(int).astype(str) + "%"

    # Check if LGD is computed
    lgd_computed = "computed_lgd" in data.columns and data["computed_lgd"].notna().all()

    # Ensure each row is a dict (not Series or string)
    data_dicts = data.to_dict(orient="records")
    paginator = Paginator(data_dicts, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Define columns to render
    columns = [
        'account_number',
        'client_name',
        'loan_type',
        'interest_rate',
        'loan_tenor',
        'client_type',
        'collateral_type',
        'computed_lgd'
    ]

    context = {
        'company': company,
        'project': project,
        'columns': columns,
        'page_obj': page_obj,
        'lgd_computed': lgd_computed
    }

    return render(request, 'impairment_engine/current_lgd.html', context)

@login_required
def compute_project_pd(request, company_slug, project_slug):
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)
    try:
        calculator = IFRS9PDCalculator()
        processor = ProjectPDProcessor(calculator)

        # Compute and update the final PDs
        processor.update_project_with_pds(project)

        # Compute the lifetime PDs
        processor.update_project_with_lifetime_pds(project)

        redirect("current_probability_given_default", company_slug=company_slug, project_slug=project_slug)
    except Exception as e:
        print(e)
        redirect("current_probability_of_default", company_slug=company_slug, project_slug=project_slug)

@login_required
def current_probability_of_default(request, company_slug, project_slug):
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    data = pd.DataFrame(project.loan_data)

    # Check if LGD is computed
    final_pd_computed = "final_pd" in data.columns and data["final_pd"].notna().all()

    # Ensure each row is a dict (not Series or string)
    data_dicts = data.to_dict(orient="records")
    paginator = Paginator(data_dicts, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Define columns to render
    columns = [
        'account_number',
        'client_name',
        'loan_type',
        'loan_tenor',
        'interest_rate',
        'model_pd',
        'final_pd'
    ]

    context = {
        'company': company,
        'project': project,
        'columns': columns,
        'page_obj': page_obj,
        'final_pd_computed': final_pd_computed
    }

    return render(request, 'impairment_engine/current_pd.html', context)


@login_required
def lifetime_probability_of_default(request, company_slug, project_slug):
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    data = pd.DataFrame(project.loan_data)

    # Check if LGD is computed
    lifetime_pd_computed = "lifetime_pd_yr1" in data.columns and data["lifetime_pd_yr1"].notna().all()

    # Ensure each row is a dict (not Series or string)
    data_dicts = data.to_dict(orient="records")
    paginator = Paginator(data_dicts, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Define columns to render
    columns = [
        'account_number',
        'client_name',
        'ltpd_yr1',
        'ltpd_yr2',
        'ltpd_yr3',
        'ltpd_yr4',
        'ltpd_yr5'
    ]

    context = {
        'company': company,
        'project': project,
        'columns': columns,
        'page_obj': page_obj,
        'lifetime_pd_computed': lifetime_pd_computed
    }

    return render(request, 'impairment_engine/lifetime_pd.html', context)


@login_required
def expected_credit_loss(request, company_slug, project_slug, stage):
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    data = pd.DataFrame(project.loan_data)
    filtered_loans = data[data['loan_stage'] == stage]
    loans_list = filtered_loans.to_dict(orient='records')

    # Check if ECL has been computed
    ecl_computed = "total_ecl" in data.columns and data["total_ecl"].notna().all()

    paginator = Paginator(loans_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Define columns to render
    columns = [
        'account_number',
        'loan_type',
        'currency',
        'loan_amount',
        'total_ecl'
    ]

    staging_title = ''
    if stage == 'stage_1':
        staging_title = 'Stage 1 Loans - Performing'
    elif stage == 'stage_2':
        staging_title = 'Stage 2 Loans - Significant Increase in Credit Risk'
    elif stage == 'stage_3':
        staging_title = 'Stage 3 Loans - Non-Performing (Defaulted)'

    context = {
        'company': company,
        'project': project,
        'columns': columns,
        'page_obj': page_obj,
        'ecl_computed': ecl_computed,
        'staging': staging_title
    }

    return render(request, 'impairment_engine/current_ecl.html', context)

@login_required
def compute_project_ecl(request, company_slug, project_slug):
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)
    try:
        ecl_processor = ProjectECLProcessor()

        # Compute and update the final ECL
        ecl_processor.update_project_with_ecls(project)

        # Update the project data
        ecl_processor.update_project_with_ecls(project)

        redirect("expected_credit_loss", company_slug=company.slug, project_slug=project.slug)
    except Exception as e:
        print(f"Error encountered whilst calculating ECL: {str(e)}")
        redirect("expected_credit_loss", company_slug=company.slug, project_slug=project.slug)


@login_required
def manage_branch_mappings(request, company_slug):
    """Manage branch mappings for a company"""
    company = get_object_or_404(Company, slug=company_slug)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to manage branch mappings.")
        return redirect('home')

    branch_mappings = company.branch_mappings.all().order_by('branch_name')

    context = {
        'company': company,
        'branch_mappings': branch_mappings,
    }

    return render(request, 'impairment/manage_branch_mappings.html', context)


@login_required
def add_branch_mapping(request, company_slug):
    """Add individual branch mapping"""
    company = get_object_or_404(Company, slug=company_slug)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to add branch mappings.")
        return redirect('home')

    if request.method == 'POST':
        form = BranchMappingForm(request.POST)
        if form.is_valid():
            branch_mapping = form.save(commit=False)
            branch_mapping.company = company
            branch_mapping.save()
            messages.success(request, "Branch mapping added successfully!")
            return HttpResponseRedirect(reverse('manage_branch_mappings', args=[company_slug]))
    else:
        form = BranchMappingForm()

    context = {
        'company': company,
        'form': form,
    }

    return render(request, 'impairment/add_branch_mapping.html', context)


@login_required
def bulk_upload_branch_mappings(request, company_slug):
    """Bulk upload branch mappings via CSV"""
    company = get_object_or_404(Company, slug=company_slug)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to upload branch mappings.")
        return redirect('home')

    if request.method == 'POST':
        form = BranchMappingBulkForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']

            try:
                # Read CSV file
                file_data = csv_file.read().decode('utf-8')
                csv_data = csv.DictReader(io.StringIO(file_data))

                created_count = 0
                error_count = 0
                errors = []

                with transaction.atomic():
                    for row_num, row in enumerate(csv_data, start=2):
                        try:
                            branch_name = row.get('branch_name', '').strip()
                            branch_code = row.get('branch_code', '').strip()
                            is_active = row.get('is_active', 'true').strip().lower() in ['true', '1', 'yes', 'y']

                            if not branch_name or not branch_code:
                                errors.append(f"Row {row_num}: Branch name and code are required")
                                error_count += 1
                                continue

                            # Check if branch code already exists
                            if BranchMapping.objects.filter(company=company, branch_code=branch_code).exists():
                                errors.append(f"Row {row_num}: Branch code '{branch_code}' already exists")
                                error_count += 1
                                continue

                            BranchMapping.objects.create(
                                company=company,
                                branch_name=branch_name,
                                branch_code=branch_code,
                                is_active=is_active
                            )
                            created_count += 1

                        except Exception as e:
                            errors.append(f"Row {row_num}: {str(e)}")
                            error_count += 1

                if created_count > 0:
                    messages.success(request, f"Successfully created {created_count} branch mappings.")

                if error_count > 0:
                    messages.warning(request, f"{error_count} rows had errors. See details below.")
                    for error in errors[:10]:  # Show first 10 errors
                        messages.error(request, error)

                return HttpResponseRedirect(reverse('manage_branch_mappings', args=[company_slug]))

            except Exception as e:
                messages.error(request, f"Error processing CSV file: {str(e)}")
    else:
        form = BranchMappingBulkForm()

    context = {
        'company': company,
        'form': form,
    }

    return render(request, 'impairment/bulk_upload_branch_mappings.html', context)


@login_required
def edit_branch_mapping(request, company_slug, mapping_id):
    """Edit individual branch mapping"""
    company = get_object_or_404(Company, slug=company_slug)
    branch_mapping = get_object_or_404(BranchMapping, id=mapping_id, company=company)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to edit branch mappings.")
        return redirect('home')

    if request.method == 'POST':
        form = BranchMappingForm(request.POST, instance=branch_mapping)
        if form.is_valid():
            form.save()
            messages.success(request, "Branch mapping updated successfully!")
            return HttpResponseRedirect(reverse('manage_branch_mappings', args=[company_slug]))
    else:
        form = BranchMappingForm(instance=branch_mapping)

    context = {
        'company': company,
        'branch_mapping': branch_mapping,
        'form': form,
    }

    return render(request, 'impairment/edit_branch_mapping.html', context)


@login_required
def manage_cbl_parameters(request, company_slug, project_slug):
    """Manage CBL parameters for a project"""
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to manage CBL parameters.")
        return redirect('home')

    cbl_parameters = project.cbl_parameters.all().order_by('loan_type', 'currency', 'risk_segment')

    context = {
        'company': company,
        'project': project,
        'cbl_parameters': cbl_parameters,
    }

    return render(request, 'impairment/manage_cbl_parameters.html', context)


@login_required
def add_cbl_parameters(request, company_slug, project_slug):
    """Add CBL parameters for a loan type/segment"""
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to add CBL parameters.")
        return redirect('home')

    if request.method == 'POST':
        form = CBLParametersForm(request.POST)
        if form.is_valid():
            cbl_params = form.save(commit=False)
            cbl_params.project = project
            cbl_params.created_by = request.user

            # Apply company defaults if not specified
            if not cbl_params.pd_floor:
                cbl_params.pd_floor = company.default_pd_floor
            if not cbl_params.lgd_floor:
                cbl_params.lgd_floor = company.default_lgd_floor
            if not cbl_params.lgd_ceiling:
                cbl_params.lgd_ceiling = company.default_lgd_ceiling

            cbl_params.save()
            messages.success(request, "CBL parameters added successfully!")
            return HttpResponseRedirect(reverse('manage_cbl_parameters', args=[company_slug, project_slug]))
    else:
        form = CBLParametersForm()
        # Pre-populate with company defaults
        form.initial.update({
            'pd_floor': company.default_pd_floor,
            'lgd_floor': company.default_lgd_floor,
            'lgd_ceiling': company.default_lgd_ceiling,
        })

    context = {
        'company': company,
        'project': project,
        'form': form,
    }

    return render(request, 'impairment/add_cbl_parameters.html', context)


@login_required
def edit_cbl_parameters(request, company_slug, project_slug, params_id):
    """Edit CBL parameters"""
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)
    cbl_params = get_object_or_404(CBLParameters, id=params_id, project=project)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to edit CBL parameters.")
        return redirect('home')

    if request.method == 'POST':
        form = CBLParametersForm(request.POST, instance=cbl_params)
        if form.is_valid():
            form.save()
            messages.success(request, "CBL parameters updated successfully!")
            return HttpResponseRedirect(reverse('manage_cbl_parameters', args=[company_slug, project_slug]))
    else:
        form = CBLParametersForm(instance=cbl_params)

    context = {
        'company': company,
        'project': project,
        'cbl_params': cbl_params,
        'form': form,
    }

    return render(request, 'impairment/edit_cbl_parameters.html', context)


@login_required
def upload_data(request, company_slug, project_slug):
    """Upload data files to a project"""
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to upload data.")
        return redirect('home')

    if request.method == 'POST':
        form = DataUploadForm(request.POST, request.FILES)
        if form.is_valid():
            data_upload = form.save(commit=False)
            data_upload.project = project
            data_upload.uploaded_by = request.user
            data_upload.save()

            # Update project flags based on upload type
            if data_upload.upload_type == 'loan_report':
                project.loan_report_uploaded = True
            elif data_upload.upload_type == 'arrears_report':
                project.arrears_report_uploaded = True

            project.status = 'data_upload'
            project.save()

            messages.success(request, f"{data_upload.get_upload_type_display()} uploaded successfully!")
            return HttpResponseRedirect(reverse('dashboard', args=[company_slug, project_slug]))
    else:
        form = DataUploadForm()

    context = {
        'company': company,
        'project': project,
        'form': form,
    }

    return render(request, 'impairment/upload_data.html', context)


@login_required
def update_company_parameters(request, company_slug):
    """Update company-level IFRS9 and CBL parameters"""
    company = get_object_or_404(Company, slug=company_slug)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to update company parameters.")
        return redirect('home')

    if request.method == 'POST':
        form = CompanyParametersUpdateForm(request.POST, instance=company)
        if form.is_valid():
            form.save()
            messages.success(request, "Company parameters updated successfully!")
            return HttpResponseRedirect(reverse('company_detail', args=[company_slug]))
    else:
        form = CompanyParametersUpdateForm(instance=company)

    context = {
        'company': company,
        'form': form,
    }

    return render(request, 'impairment/update_company_parameters.html', context)


@login_required
def data_uploads_list(request, company_slug, project_slug):
    """List all data uploads for a project"""
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to view data uploads.")
        return redirect('home')

    uploads_list = project.data_uploads.all().order_by('-uploaded_at')
    paginator = Paginator(uploads_list, 20)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'company': company,
        'project': project,
        'page_obj': page_obj,
    }

    return render(request, 'impairment/data_uploads_list.html', context)


@login_required
@require_http_methods(["POST"])
def delete_branch_mapping(request, company_slug, mapping_id):
    """Delete a branch mapping"""
    company = get_object_or_404(Company, slug=company_slug)
    branch_mapping = get_object_or_404(BranchMapping, id=mapping_id, company=company)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to delete branch mappings.")
        return redirect('home')

    branch_mapping.delete()
    messages.success(request, "Branch mapping deleted successfully!")
    return HttpResponseRedirect(reverse('manage_branch_mappings', args=[company_slug]))


@login_required
@require_http_methods(["POST"])
def delete_cbl_parameters(request, company_slug, project_slug, params_id):
    """Delete CBL parameters"""
    company = get_object_or_404(Company, slug=company_slug)
    project = get_object_or_404(Project, slug=project_slug, company=company)
    cbl_params = get_object_or_404(CBLParameters, id=params_id, project=project)

    # Check permissions
    if not request.user.is_superuser and company.created_by != request.user:
        messages.error(request, "You don't have permission to delete CBL parameters.")
        return redirect('home')

    loan_type = cbl_params.loan_type
    cbl_params.delete()
    messages.success(request, f"CBL parameters for {loan_type} deleted successfully!")
    return HttpResponseRedirect(reverse('manage_cbl_parameters', args=[company_slug, project_slug]))


# @login_required
# def project_dashboard(request, company_slug, project_slug):
#     """Project dashboard with processing status and statistics"""
#     company = get_object_or_404(Company, slug=company_slug)
#     project = get_object_or_404(Project, slug=project_slug, company=company)
#
#     # Check permissions
#     if not request.user.is_superuser and company.created_by != request.user:
#         messages.error(request, "You don't have permission to view this project dashboard.")
#         return redirect('home')
#
#     # Get dashboard statistics
#     loan_accounts_count = project.loan_accounts.count()
#     arrears_accounts_count = project.arrears_accounts.count()
#
#     # Get staging distribution
#     stage_1_count = IFRS9Stage.objects.filter(
#         loan_account__project=project,
#         current_stage='stage_1'
#     ).count()
#     stage_2_count = IFRS9Stage.objects.filter(
#         loan_account__project=project,
#         current_stage='stage_2'
#     ).count()
#     stage_3_count = IFRS9Stage.objects.filter(
#         loan_account__project=project,
#         current_stage='stage_3'
#     ).count()
#
#     # Get ECL summary
#     total_ecl = ECLCalculation.objects.filter(
#         loan_account__project=project
#     ).aggregate(
#         total=models.Sum('final_ecl')
#     )['total'] or 0
#
#     context = {
#         'company': company,
#         'project': project,
#         'loan_accounts_count': loan_accounts_count,
#         'arrears_accounts_count': arrears_accounts_count,
#         'stage_1_count': stage_1_count,
#         'stage_2_count': stage_2_count,
#         'stage_3_count': stage_3_count,
#         'total_ecl': total_ecl,
#         'recent_uploads': project.data_uploads.all().order_by('-uploaded_at')[:3],
#     }
#
#     return render(request, 'impairment/project_dashboard.html', context)